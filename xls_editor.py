import openpyxl
workbook = openpyxl.load_workbook("example.xlsx")
sheet = workbook.active
# или
sheet = workbook["Sheet1"]
cell_value = sheet["A1"].value
cell_value = sheet.cell(row=1, column=1).value
for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
    print(row)

